import openpyxl
import os
import shutil
from datetime import datetime

def generate_voucher_excel(data, template_path=None):
    if not template_path:
        template_path = r"C:\Users\baewoong.kim\Desktop\고려제강(2025).xlsx"

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Voucher 템플릿 엑셀 파일을 찾을 수 없습니다: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    ws['W6'] = data.get('pr_no', '')
    ws['D9'] = data.get('pr_title', '')
    ws['W9'] = data.get('amount', 0)
    ws['W16'] = data.get('vat', 0)
    ws['W17'] = data.get('total_amount', 0)
    ws['A6'] = data.get('date', '')
    ws['J6'] = data.get('supplier', '')
    
    ws['P32'] = data.get('total_amount', 0)
    ws['P33'] = data.get('amount', 0)
    ws['V34'] = data.get('vat', 0)

    pr_no = data.get('pr_no', 'NO_PR')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"Voucher_{pr_no}_{timestamp}.xlsx"
    desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    out_path = os.path.join(desktop, out_name)
    wb.save(out_path)
    return out_path

def archive_voucher_package(data, pdf_paths, excel_path=None):
    """
    [건별/업체별 자동 폴더 생성 보관소]
    날짜_거래처명_PR번호 폴더 생성 후 PDF 4종 및 엑셀 파일 자동 정돈 보관
    """
    date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    supplier_str = data.get('supplier', '거래처')
    pr_no_str = data.get('pr_no', 'PR')

    folder_name = f"{date_str}_{supplier_str}_{pr_no_str}"
    base_vault = os.path.join(os.path.expanduser('~'), 'Desktop', 'Voucher_보관소')
    target_dir = os.path.join(base_vault, folder_name)
    os.makedirs(target_dir, exist_ok=True)

    archived_files = []

    # PDF 서류 복사 보관
    for p in pdf_paths:
        if p and os.path.exists(p):
            fname = os.path.basename(p)
            dest = os.path.join(target_dir, fname)
            shutil.copy2(p, dest)
            archived_files.append(dest)

    # 엑셀 파일 보관
    if excel_path and os.path.exists(excel_path):
        fname = os.path.basename(excel_path)
        dest = os.path.join(target_dir, fname)
        shutil.copy2(excel_path, dest)
        archived_files.append(dest)

    return target_dir, archived_files

if __name__ == '__main__':
    print("excel_handler with auto-archiver ready")
